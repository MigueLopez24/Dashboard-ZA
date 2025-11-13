import pandas as pd
from datetime import datetime
import tempfile
import os
from src.data.sql_connector import obtener_historial_fallas

def generar_dataframe_fallas(dias: int = 7):
    """
    Obtiene el historial de fallas y lo convierte en un DataFrame de pandas.
    """
    resultado = obtener_historial_fallas(dias)
    if "error" in resultado:
        raise Exception(resultado["error"])
    df = pd.DataFrame(resultado["data"])
    if not df.empty:
        # Formatear fechas y duración
        df['fecha_hora_inicio'] = pd.to_datetime(df['fecha_hora_inicio'])
        df['fecha_hora_fin'] = pd.to_datetime(df['fecha_hora_fin'])
        df['duracion'] = df['duracion_minutos'].apply(lambda x: f"{x//60}h {x%60}m" if pd.notnull(x) else "")
    return df

def exportar_fallas_excel(dias: int = 7, filepath=None):
    
    df = generar_dataframe_fallas(dias)
    if filepath is None:
        fecha = datetime.now().strftime("%Y%m%d")
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, f"reporte_fallas_{fecha}.xlsx")

    # Validación para DataFrame vacío
    if df.empty:
        # Define columnas esperadas para evitar KeyError
        columnas = [
            'nombre_dispositivo', 'tipo_dispositivo', 'fecha_hora_inicio',
            'fecha_hora_fin', 'duracion_minutos', 'estado_falla', 'duracion'
        ]
        df = pd.DataFrame(columns=columnas)
        resumen = {
            "Total de fallas": [0],
            "Total dispositivos afectados": [0],
            "Total horas caídas": [0]
        }
    else:
        resumen = {
            "Total de fallas": [len(df)],
            "Total dispositivos afectados": [df['nombre_dispositivo'].nunique()],
            "Total horas caídas": [df['duracion_minutos'].sum() // 60]
        }

    # --- Escribir Excel con formato ---
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Fallas", startrow=4)
        pd.DataFrame(resumen).to_excel(writer, index=False, sheet_name="Resumen")

        workbook = writer.book
        ws = writer.sheets["Fallas"]

        # Insertar logo (asegura ruta absoluta)
        try:
            from openpyxl.drawing.image import Image as XLImage
            logo_paths = [
                os.path.join(os.getcwd(), "assets", "logo_zavala_aguilar.png"),
                os.path.join(os.getcwd(), "assets", "logo_zavala_aguilar.PNG"),
                os.path.join(os.getcwd(), "assets", "icons", "logo_zavala_aguilar.png"),
            ]
            logo_path = next((p for p in logo_paths if os.path.exists(p)), None)
            if logo_path:
                img = XLImage(logo_path)
                img.height = 60
                img.width = 180
                ws.add_image(img, "A1")
        except Exception as e:
            pass

        # Título
        ws.merge_cells('A2:F2')
        ws['A2'] = "Reporte de Fallas por Dispositivo"
        ws['A2'].font = ws['A2'].font.copy(bold=True, size=16)
        ws['A2'].alignment = ws['A2'].alignment.copy(horizontal="center", vertical="center")

        # Ajustar ancho de columnas automáticamente
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        # Estilo azul para encabezados y filas alternas, y centrar todo el contenido
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[5]:  # Encabezados (row=5 por startrow=4)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Filas alternas azules claras y centrar datos
        fill_even = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
        for i, row in enumerate(ws.iter_rows(min_row=6, max_row=ws.max_row), start=0):
            for cell in row:
                cell.alignment = center_alignment
            if i % 2 == 0:
                for cell in row:
                    cell.fill = fill_even

    return filepath

def exportar_fallas_pdf(dias: int = 7, filepath=None):
    """
    Exporta el historial de fallas a un archivo PDF (simple, tabla).
    """
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    df = generar_dataframe_fallas(dias)
    if filepath is None:
        fecha = datetime.now().strftime("%Y%m%d")
        filepath = f"reporte_fallas_{fecha}.pdf"
    resumen = [
        f"Total de fallas: {len(df)}",
        f"Total dispositivos afectados: {df['nombre_dispositivo'].nunique()}",
        f"Total horas caídas: {df['duracion_minutos'].sum() // 60}"
    ]
    with PdfPages(filepath) as pdf:
        # Resumen
        plt.figure(figsize=(8, 1))
        plt.axis('off')
        plt.title("Resumen de Fallas (últimos 7 días)", fontsize=14)
        for i, line in enumerate(resumen):
            plt.text(0, 0.8 - i*0.3, line, fontsize=12)
        pdf.savefig(); plt.close()

        # Tabla de fallas (en varias páginas si es necesario)
        if not df.empty:
            cols = ['nombre_dispositivo', 'tipo_dispositivo', 'fecha_hora_inicio', 'fecha_hora_fin', 'duracion', 'estado_falla']
            rows = df[cols].astype(str).values.tolist()
            header = ['Dispositivo', 'Tipo', 'Inicio', 'Fin', 'Duración', 'Estado']
            for i in range(0, len(rows), 30):
                plt.figure(figsize=(12, 8))
                plt.axis('off')
                plt.title("Historial de Fallas", fontsize=14)
                table = plt.table(cellText=rows[i:i+30], colLabels=header, loc='center', cellLoc='center')
                table.auto_set_font_size(False)
                table.set_fontsize(10)
                table.scale(1, 1.5)
                pdf.savefig(); plt.close()
    return filepath
